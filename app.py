from flask import Flask, render_template, request
import openpyxl
from flaskwebgui import FlaskUI

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/result', methods=["GET","POST"])
def result():
    if request.method == 'POST':
        column = request.form.get("columnidentifier")
        row1 = request.form.get("head")
        rowend = request.form.get("foot")
        multiplier = request.form.get("multiplier")
        column.strip()
        row1.strip()
        rowend.strip()
        multiplier.strip()
        column.upper()
        rowend = int(rowend)
        row1 = int(row1)
        rowend += 1
        multiplier = float(multiplier)
        multiplier = 1 + multiplier/100
        wb = openpyxl.load_workbook("input.xlsx")
        ws = wb.active
        for i in range(row1,rowend):
            ws["{col}{digit}".format(col = column, digit = i)].value = (ws["{col}{digit}".format(col = column, digit = i)].value*multiplier)
            print(ws["{col}{digit}".format(col = column, digit = i)].value)
        wb.save("output.xlsx")
    
    return render_template("result.html")

